import openpyxl

def make_DB():
    excel = openpyxl.load_workbook('bugs.xlsx')
    wb = excel.get_sheet_by_name('bugs')
    excel = openpyxl.load_workbook('genie.xlsx')
    wg = excel.get_sheet_by_name('genie')
    excel = openpyxl.load_workbook('melon.xlsx')
    wm = excel.get_sheet_by_name('melon')
    excel = openpyxl.load_workbook('naver.xlsx')
    wn = excel.get_sheet_by_name('naver')
    DB = [] # 이름, 가수, Bugs, Genie, Melon, Naver, 순위(B), 순위(G), 순위(M), 순위(N), 점수

    for i in range(0,100) :
        title = wb['A'][i].value
        singer = wb['B'][i].value
        songid = wb['C'][i].value
        singer = singer.upper()
        if(singer.find("(")):
            singer = singer.split("(")[0]
        DB.append([title,singer,songid,'0','0','0',(100-i),'0','0','0','0'])

    for i in range(0,100) :
        flag = 1
        title = wg['A'][i].value
        singer = wg['B'][i].value
        songid = wg['C'][i].value
        for E in DB:
            if title in E:
                if(E[1].find("(")):
                    E[1] = E[1].split("(")[0]
                if(singer.find("(")):
                    singer = singer.split("(")[0]
                E[1] = E[1].upper()
                singer = singer.upper()
                if (singer.replace(" ", "") == E[1].replace(" ", "")):
                    E[3] = songid
                    E[7] = (100-i)
                    flag = 0
        if(flag==1):
            DB.append([title,singer,'0',songid,'0','0','0',(100-i),'0','0','0'])

    for i in range(0,100) :
        flag=1
        title = wm['A'][i].value
        singer = wm['B'][i].value
        songid = wm['C'][i].value
        for E in DB:
            if title in E:
                if(E[1].find("(")):
                    E[1] = E[1].split("(")[0]
                if(singer.find("(")):
                    singer = singer.split("(")[0]
                E[1] = E[1].upper()
                singer = singer.upper()
                if (singer.replace(" ", "") == E[1].replace(" ", "")):
                    E[4] = songid
                    E[8] = (100-i)
                    flag = 0
        if(flag==1):
            DB.append([title,singer,'0','0',songid,'0','0','0',(100-i),'0','0'])
    for i in range(0,100) :
        flag=1
        title = wn['A'][i].value
        singer = wn['B'][i].value
        songid = wn['C'][i].value
        for E in DB:
            if title in E:
                if(E[1].find("(")):
                    E[1] = E[1].split("(")[0]
                if(singer.find("(")):
                    singer = singer.split("(")[0]
                E[1] = E[1].upper()
                singer = singer.upper()
                if (singer.replace(" ", "") == E[1].replace(" ", "")):
                    E[5] = songid
                    E[9] = (100-i)
                    flag = 0
        if(flag==1):
            DB.append([title,singer,'0','0','0',songid,'0','0','0',(100-i),'0'])
    return DB

def save() :
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0,title="DB")
    DB_list = make_DB()
    for i in range(0,len(DB_list)):
        DB_list[i][10] = str(int(DB_list[i][6])*0.2+int(DB_list[i][7])*0.3+int(DB_list[i][8])*0.4+int(DB_list[i][9])*0.1) 
    r = 1
    for i in range(1,len(DB_list)):
        for j in range(1,12):
            ws.cell(row=i, column=j).value = DB_list[i-1][j-1]
    wb.save('DB.xlsx')

if __name__ == "__main__":
    save()
